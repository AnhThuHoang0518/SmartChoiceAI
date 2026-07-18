# -*- coding: utf-8 -*-
"""Nap cac sheet DIEN TU (tablet, dong ho, man hinh, PC, may in) -> processed/*.csv

Chay:  python scripts/nap_dmx_dien_tu.py

Nhom nay khac gia dung: khong co 'so nguoi su dung' - rang buoc la thong so
truc tiep (inch, GB, mAh...). Muc 1: loc theo rang buoc khach neu + cham theo
uu tien do duoc tu cot. Bang map MUC DICH -> nguong (vd 'cho con hoc' -> may
inch/GB) KHONG lam o day vi khong co trong du lieu - can nghiep vu chot,
khong tu che (xem ghi chu cuoi file).
"""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from backend.app.services.parse_dmx import la_rong, parse_gia, parse_qua  # noqa: E402

GOC = Path("data/raw/Spec_cate_gia.xlsx")


def _num(s):
    if la_rong(s):
        return None
    m = re.search(r"([\d.]+)", str(s).replace(",", "."))
    return float(m.group(1)) if m else None


def p_inch(s):
    return _num(s)          # '8.7"' / '27 inch' / '1.3 inch'


def p_gb(s):
    """'64 GB' -> 64 | '1 TB' -> 1024."""
    if la_rong(s):
        return None
    t = str(s).lower()
    v = _num(t)
    if v is None:
        return None
    return v * 1024 if "tb" in t else v


def p_ngay(s):
    """'Khoang 4 ngay' / 'Khoang 5 ngay (GPS) | 14 ngay (tiet kiem)' -> MAX ngay.

    Lay max co chu y: hang cong bo nhieu che do, con so to nhat la che do tiet
    kiem - van la so HANG cong bo, khong phai minh doan."""
    if la_rong(s):
        return None
    so = re.findall(r"([\d.]+)\s*ngày", str(s).lower())
    return max(float(x) for x in so) if so else None


def p_max_so(s):
    """'250-2000 trang/thang' -> 2000 | '99% sRGB | 72% NTSC' -> 99 (max)."""
    if la_rong(s):
        return None
    so = re.findall(r"[\d.]+", str(s).replace(",", "."))
    return max(float(x) for x in so) if so else None


def p_phan_giai(s):
    """Chuan hoa do phan giai ve nhan gon: 4K/2K/Full HD/HD."""
    if la_rong(s):
        return ""
    t = str(s).lower()
    for tu, nhan in [("4k", "4K"), ("uhd", "4K"), ("3840", "4K"),
                     ("2k", "2K"), ("qhd", "2K"), ("2560", "2K"),
                     ("full hd", "Full HD"), ("1920", "Full HD"), ("hd", "HD")]:
        if tu in t:
            return nhan
    return str(s)[:20]


SPEC = {
    "may_tinh_bang": {
        "sheet": "Máy tính bảng",
        "cot": {
            "man_inch": ("Kích thước màn hình", p_inch),
            "ram_gb": ("RAM", p_gb),
            "pin_mah": ("Dung lượng pin", _num),
            "luu_tru_gb": ("Dung lượng lưu trữ", p_gb),
            "nang_g": ("Khối lượng máy", _num),
        },
        "chu": {"sim": "SIM"},
        "bat_buoc": ["man_inch"],
    },
    "dong_ho": {
        "sheet": "Đồng hồ thông minh",
        "cot": {
            "pin_ngay": ("Thời gian sử dụng", p_ngay),
            "pin_mah": ("Dung lượng pin", _num),
            "man_inch": ("Kích thước màn hình", p_inch),
        },
        "chu": {"nghe_goi": "Thực hiện cuộc gọi", "khang_nuoc": "Chuẩn chống nước, bụi"},
        "bat_buoc": [],
    },
    "man_hinh": {
        "sheet": "Màn hình máy tính",
        "cot": {
            "man_inch": ("Kích thước màn hình", p_inch),
            "dap_ung_ms": ("Thời gian đáp ứng", _num),
            "do_sang_nit": ("Độ sáng", _num),
            "loa_co": ("Loa", lambda v: 1.0 if str(v or "").strip().lower() == "có" else None),
            "phu_mau_pct": ("Độ phủ màu", p_max_so),
        },
        "chu": {"phan_giai": ("Độ phân giải", p_phan_giai), "tam_nen": "Tấm nền"},
        "bat_buoc": ["man_inch"],
    },
    "may_tinh_ban": {
        "sheet": "Máy tính để bàn",
        "cot": {
            "ram_gb": ("RAM", p_gb),
            "ssd_gb": ("Ổ cứng", p_gb),
        },
        "chu": {"cpu": "Loại CPU", "cong_nghe_cpu": "Công nghệ CPU"},
        "bat_buoc": [],
    },
    "may_in": {
        "sheet": "Máy in",
        "cot": {
            "toc_do_trang": ("Tốc độ in", _num),
            "khay_to": ("Khay nạp giấy", _num),
            "cong_suat_thang": ("Công suất theo nghiệp vụ", p_max_so),
        },
        "chu": {"loai": "Loại sản phẩm", "ket_noi": "Kết nối"},
        "bat_buoc": [],
    },
}


def nap_mot_nganh(wb, ten: str, spec: dict) -> None:
    ws = wb[spec["sheet"]]
    rows = ws.iter_rows(values_only=True)
    hdr = list(next(rows))
    idx = {h: i for i, h in enumerate(hdr)}

    def g(r, cot):
        i = idx.get(cot)
        return r[i] if i is not None and i < len(r) else None

    ra, tong = [], 0
    for r in rows:
        if not any(r):
            continue
        tong += 1
        gia, gia_goc = parse_gia(g(r, "giá gốc"), g(r, "giá khuyến mãi"))
        if not gia:
            continue
        dong = {
            "ma_sp": str(g(r, "sku") or g(r, "model_code") or "").strip(),
            "ten": f"{str(g(r, 'brand') or '').strip()} {g(r, 'model_code')}".strip(),
            "hang": str(g(r, "brand") or "").strip(),
            "gia": gia, "gia_goc": gia_goc or gia,
            "qua": parse_qua(g(r, "khuyến mãi quà")),
        }
        for c, (cot_goc, ham) in spec["cot"].items():
            v = ham(g(r, cot_goc))
            dong[c] = "" if v is None else v
        for c, meta in spec["chu"].items():
            if isinstance(meta, tuple):
                cot_goc, ham = meta
                dong[c] = ham(g(r, cot_goc))
            else:
                dong[c] = str(g(r, meta) or "").strip()
        if any(dong[c] == "" for c in spec["bat_buoc"]):
            continue
        ra.append(dong)

    out = Path(f"data/processed/{ten}.csv")
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(ra[0].keys()))
        w.writeheader()
        w.writerows(ra)
    print(f"{ten:14s} {tong:5d} dong -> ban duoc {len(ra):4d} -> {out}")


def p_gio(s):
    """'8 - 10 tieng' -> 8 (can duoi bao thu cua dai hang cong bo).

    Dung 10 de loc dieu kien "it nhat 9 gio" se xac nhan qua tay trong khi
    chinh nguon noi thoi luong co the chi 8 gio. Ranking va hard filter dung
    can duoi de khong hua qua muc toi thieu.
    """
    if la_rong(s):
        return None
    so = re.findall(r"[\d.]+", str(s))
    return min(float(x) for x in so) if so else None


def nap_micro(wb) -> None:
    """GOP 2 sheet micro thanh 1 nganh: 'Micro karaoke' chi co 5 SKU co gia -
    tach 2 nganh rieng la vo nghia voi khach."""
    ra = []
    for sheet, nhom in [("Micro karaoke", "karaoke"), ("Micro thu âm điện thoại", "thu âm")]:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        hdr = list(next(rows))
        idx = {h: i for i, h in enumerate(hdr)}

        def g(r, cot):
            i = idx.get(cot)
            return r[i] if i is not None and i < len(r) else None

        for r in rows:
            if not any(r):
                continue
            gia, gia_goc = parse_gia(g(r, "giá gốc"), g(r, "giá khuyến mãi"))
            if not gia:
                continue
            ra.append({
                "ma_sp": str(g(r, "sku") or g(r, "model_code") or "").strip(),
                "ten": f"{str(g(r, 'brand') or '').strip()} {g(r, 'model_code')}".strip(),
                "hang": str(g(r, "brand") or "").strip(),
                "gia": gia, "gia_goc": gia_goc or gia,
                "khoang_cach_m": _num(g(r, "Khoảng cách truyền")) or "",
                "pin_gio": p_gio(g(r, "Thời gian sử dụng")) or "",
                "nhom": nhom,
                "loai": str(g(r, "Loại sản phẩm") or "").strip(),
                "ket_noi": str(g(r, "Kết nối") or "").strip(),
                "qua": parse_qua(g(r, "khuyến mãi quà")),
            })
    out = Path("data/processed/micro.csv")
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(ra[0].keys()))
        w.writeheader()
        w.writerows(ra)
    print(f"{'micro':14s} (2 sheet gop) -> ban duoc {len(ra):4d} -> {out}")


def main() -> None:
    wb = openpyxl.load_workbook(GOC, data_only=True, read_only=True)
    for ten, spec in SPEC.items():
        nap_mot_nganh(wb, ten, spec)
    nap_micro(wb)


# ── GHI CHU cho buoc sau (can nghiep vu, KHONG tu che) ──────────────────────
# Bang map MUC DICH -> nguong toi thieu, vd:
#   "cho con hoc online"  -> tablet: man >= ? inch, RAM >= ? GB
#   "choi game"           -> man hinh: dap ung <= ? ms; PC: RAM >= ? GB
#   "chay bo/the thao"    -> dong ho: pin >= ? ngay, khang nuoc >= IP?
# Cac nguong nay phai do NGUOI HIEU BAN HANG chot roi bo vao config
# (muc "muc_dich" du kien) - de AI tu dat nguong la vi pham luat khong bia.

if __name__ == "__main__":
    main()
