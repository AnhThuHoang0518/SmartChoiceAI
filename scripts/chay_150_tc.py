# -*- coding: utf-8 -*-
"""Chay 150 test case VIAC qua chatbot -> xuat ket qua ra xlsx.

Chay:  python scripts/chay_150_tc.py            (mac dinh: khong LLM, nhanh)
       python scripts/chay_150_tc.py --llm      (LLM that - cham, ton credit)

Sua 3 loi cua lan chay truoc (ban do team lam):
  1. MOI test case MOT phien_id MOI - khong de trang thai cau truoc lay sang
     cau sau (bang cu: "23.4m²" lay khap noi vi dung chung 1 phien).
  2. Chay tren DU LIEU THAT neu co (data/processed/) - khong phai catalog mau
     Cosmo/Alpha vai chuc may.
  3. Khi bot hoi nguoc thi tra loi mac dinh CO BIEN THIEN theo Test ID - cac
     case khac nhau khong bi dong khung mot bo loc giong het nhau.
"""
from __future__ import annotations

import os
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

if "--llm" in sys.argv:
    from dotenv import load_dotenv
    load_dotenv()
else:
    os.environ["LLM_NHA_CUNG_CAP"] = "luat"

import openpyxl  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402

from backend.app.main import app  # noqa: E402

NGUON = Path("evaluations/datasets/testcase_viac_150.xlsx")
RA = Path("evaluations/results/ket_qua_150_tc.xlsx")

# Tra loi khi bot hoi nguoc - BIEN THIEN theo so thu tu case de cac case
# khong nêu ngân sách/diện tích không bị chung một bộ lọc y hệt nhau.
TEN_NGANH_NGU_CANH = {
    "tủ lạnh": "tủ lạnh",
    "máy lạnh": "máy lạnh",
    "máy giặt": "máy giặt",
    "máy sấy": "máy sấy",
    "máy rửa chén": "máy rửa chén",
    "tủ đông/tủ mát": "tủ đông",
    "máy nước nóng": "máy nước nóng",
    "micro karaoke": "micro karaoke",
    "micro thu âm điện thoại": "micro thu âm",
    "đồng hồ thông minh": "đồng hồ thông minh",
    "máy tính để bàn": "máy tính để bàn",
    "màn hình máy tính": "màn hình máy tính",
    "máy in": "máy in",
    "máy tính bảng": "máy tính bảng",
    "tivi": "tivi",
    "laptop": "laptop",
    "điện thoại": "điện thoại",
}


def nganh_ngu_canh(category: str) -> str:
    """Category la ngu canh cua testcase, tuong duong khach dang o trang nganh.

    Neu bo cot nay, cac cau nhu 'pin co du 6 gio khong?' khong the tu tiet lo
    la dang noi micro hay dien thoai; harness cu tra loi 'may lanh' va cham sai
    router thay vi cham logic cua nganh trong testcase.
    """
    ten = str(category or "").strip().lower()
    return TEN_NGANH_NGU_CANH.get(ten, str(category or "").strip())


def tra_loi_mac_dinh(i: int, cau_hoi: str, category: str) -> str:
    t = cau_hoi.lower()
    if "sản phẩm nào" in t or "sản phẩm gì" in t:
        return nganh_ngu_canh(category) or "máy lạnh"
    if "bao nhiêu m²" in t or "diện tích" in t:
        return f"phòng {14 + (i % 5) * 4}m2"          # 14/18/22/26/30
    if "nắng" in t:
        return "có nắng" if i % 2 else "không nắng"
    if "phòng ngủ" in t:
        return "phòng ngủ" if i % 2 else "phòng khách"
    if "mấy người" in t or "người dùng" in t:
        return f"{2 + i % 4} người"                    # 2-5 nguoi
    if "bao nhiêu kg" in t:
        return f"{7 + (i % 3) * 2} kg"
    if "mấy bữa" in t:
        return f"{1 + i % 3} bữa"
    if "khu vực" in t or "tỉnh" in t:
        return ["Hà Nội", "Đà Nẵng", "TP HCM"][i % 3]
    # ngan sach xoay vong - khong co muc <8tr: tu lanh/may lanh duoi 5tr gan
    # nhu khong ton tai, de muc do la tu ep bot vao 'khong_co_may' oan
    return f"khoảng {[10, 12, 15, 20, 25][i % 5]} triệu"


def main() -> None:
    co_that = Path("data/processed/may_lanh.csv").exists()
    print(f"Du lieu: {'THAT (data/processed)' if co_that else 'MAU (Cosmo/Alpha) - ket qua se don dieu!'}")
    print(f"LLM    : {os.getenv('LLM_NHA_CUNG_CAP', 'theo .env')}")

    c = TestClient(app)
    wb = openpyxl.load_workbook(NGUON)
    ws = wb.active
    hdr = [cell.value for cell in ws[1]]
    i_query = hdr.index("User query / Scenario")
    i_category = hdr.index("Category")

    # ghi ket qua vao 3 cot cuoi
    n_cot = len(hdr)
    ws.cell(row=1, column=n_cot + 1, value="loai_cuoi")
    ws.cell(row=1, column=n_cot + 2, value="tra_loi_cuoi")
    ws.cell(row=1, column=n_cot + 3, value="so_cau_hoi_nguoc")

    dem = {"tong": 0}
    for i, row in enumerate(ws.iter_rows(min_row=2), start=0):
        cau = row[i_query].value
        if not cau or not str(cau).strip():
            continue
        dem["tong"] += 1

        # MOI case MOT phien - trang thai khong lay
        pid = None
        category = str(row[i_category].value or "")
        # Cot Category la ngu canh co san cua testcase (tuong duong trang nganh
        # tren website). Gui kem ngu canh nay de khong bien cau mo ho thanh may
        # lanh bang cau tra loi mac dinh cua harness.
        ten_nganh = nganh_ngu_canh(category)
        cau_gui = f"{ten_nganh}: {cau}" if ten_nganh else str(cau)
        r = c.post("/api/chat", json={"tin_nhan": cau_gui}).json()
        pid = r["phien_id"]
        so_hoi = 0
        while r["loai"] in {"cau_hoi", "xac_nhan_nganh"} and so_hoi < 5:
            so_hoi += 1
            dap = ten_nganh if r["loai"] == "xac_nhan_nganh" \
                else tra_loi_mac_dinh(i, r["text"], category)
            r = c.post("/api/chat", json={"tin_nhan": dap, "phien_id": pid}).json()

        rn = row[0].row
        ws.cell(row=rn, column=n_cot + 1, value=r["loai"])
        ws.cell(row=rn, column=n_cot + 2, value=(r["text"] or "")[:500])
        ws.cell(row=rn, column=n_cot + 3, value=so_hoi)
        dem[r["loai"]] = dem.get(r["loai"], 0) + 1
        if dem["tong"] % 25 == 0:
            print(f"  ... {dem['tong']} case")

    RA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(RA)
    print(f"\nXong {dem['tong']} case -> {RA}")
    for k, v in sorted(dem.items()):
        if k != "tong":
            print(f"  {k:22s} {v}")


if __name__ == "__main__":
    main()
